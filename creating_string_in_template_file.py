import openpyxl #Подключаем библиотеку Openpyxl
def creating_string_in_template_file(r):
    #берем из DB.xlsx сроку что бы создать запись в файле template.xlsx
    book_DB = openpyxl.load_workbook("DB.xlsx", read_only=True)
    sheet_DB = book_DB.active
    kod = sheet_DB.cell(row=r, column=2).value
    #artikul = sheet_DB.cell(row=r, column=3).value
    naimenovanie = sheet_DB.cell(row=r, column=4).value
    book_template = openpyxl.load_workbook("template.xlsx") #Открываем тестовый Excel файл
    sheet_template = book_template.active
    #worksheet = book['Sheet1'] #Делаем его активным
    sheet_template[coordinates_kod].value= kod  #
    sheet_template[coordinates_artikul].value= artikul  #
    sheet_template[coordinates_naimenovanie].value= naimenovanie  #
    book_template.save("template.xlsx") #Сохраняем измененный файл
    book_template.close()
