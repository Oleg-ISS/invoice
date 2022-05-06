import openpyxl 
import re
def creating_template():
    #берем название изделия
    #извлекаем содержимое ячейки sheet.cell(row=r, column=c).value
    r=1
    c=1
    book_invoice = openpyxl.load_workbook("invoice.xlsx", read_only=True)
    sheet_invoice = book_invoice.active
    print(sheet.cell(row=r, column=c).value)
    #По очереди пытаемся заменить название изделия с поставщика на наше, если получиться то меняем если нет то остается как было
    item = re.sub('O-ring','Кольцо круглого сечения',sheet_invoice.cell(row=r, column=c).value)
    item = re.sub('Oil Seal','Манжета армированная',sheet_invoice.cell(row=r, column=c).value)
    print(item)
    #
    #
    #
    #берем размер
    #берем из ячейки sheet.cell(row=r, column=c+1).value содержимое, это строковая переменная

    size = sheet_invoice.cell(row=r, column=c+1).value
    print('Содержимое ячейки ',size)
    print(type(size))
    #меняем в этой переменной символ "x" на символ "\D"
    #!!!!! символ х из кирилической раскладки
    x = 'х'
    print('Меняем символ из кирилицы "х" на "\D" ',re.sub(x,r'\\D',size))
    print('Выполним проверку на соответствие с помощию match = re.search(size, b)')
    match = re.search(re.sub(x,r'\\D',size), size)
    print(match[0] if match else 'Not found')
    #
    #
    #
    #
    #меняем в этой переменной символ "*" на символ "\D"
    x = '\*'
    print('Меняем символ из кирилицы "*" на "\D" ',re.sub(x,r'\\D',size))
    size = re.sub(x,r'\\D',size)
    #print('Выполним проверку на соответствие с помощию match = re.search(size, b)')
    #match = re.search(re.sub(x,r'\\D',size), size)
    #print(match[0] if match else 'Not found')
    #
    #
    #
    #берем материал
    material = sheet_invoice.cell(row=r, column=c+2).value
    #
    #
    #объединяем название товара и шаблон размера в одну строку
    #Close the workbook after reading
    book.close()
    return item + ' ' + size + ' ' + material
print(creating_template())
