import openpyxl 
import re
#обьявляем функции
#открываем файлы
r=1
c=1
i_main = 0 #счетчик итераций для фукции searching()
#открываем таблицу invoice.xlsx
book_invoice = openpyxl.load_workbook("invoice.xlsx")#открываем файл invoice.xlsx на чтение и запись
sheet_invoice = book.active
#открываем таблицу DB.xlsx для поиска в ней нужной строки
book_DB = openpyxl.load_workbook("DB.xlsx", read_only=true)
sheet_DB = book.active

for row in range(1,sheet_invoice.max_row+1):#получаем количество строк на рабочем листе +1 чтобы учесть последнюю строку
    
    #обнуляем new_string
    new_string =0
    nazvanie = sheet_invoice.cell(row=r, column=c).value
    razmer = sheet_invoice.cell(row=r, column=c+1).value
    material = sheet_invoice.cell(row=r, column=c+2).value
    #создаем шаблон-строку для поиска по БД на основе файла invoice.xlsx 
    template = creating_template(nazvanie,razmer,material)
    #запускаем функцию поиска строки, которая соответствует шаблону-строке в файле DB.xlsx
    new_string = searching(template)
    #открываем на запись файл-шаблон для импорта счета в мой склад
    #запускаем функцию которая создаст шаблон из строки инвойса
    #передадим шаблон в функцию поиска в базе мой склад
    #если в базе есть строка соответствующая шаблону запустим функцию записи в файл-шаблон мой склад
    if new_string != 0:
        creating_string_in_template_file(new_string)
    else:
        #если строка не найдена запустим функцию записи состоянии не нашел
        creating_dont_search_string(i) 
print("Работа программы завершена")
